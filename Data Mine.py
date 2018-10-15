from openpyxl import *
from statistics import *
from textblob import TextBlob

#Load workbooks and active sheet
exbook=load_workbook('restReviews.xlsx')
exsheet1=exbook.active

#Name the new sheet
restBook=Workbook()
sheet1=restBook.active
sheet1.title='Data'

#Create List for star ratings and sentiment
bksentimentReviewList=[]
bkratingList=[]
subsentimentReviewList=[]
subratingList=[]

#Create for function 
for i in range (2,21):
    #append rating to List
    restRating=int(exsheet1.cell(column=1 , row=i).value)
    #append text reviews to List
    restReviewText=exsheet1.cell(column=2 , row=i).value
    sentRating=(TextBlob(restReviewText).sentiment.polarity)
    #Create if functions to seperate the burger kings and subway names
    restName=exsheet1.cell(column=3 , row=i).value
    if restName=="Burger King":
        bkratingList.append(restRating)
        bksentimentReviewList.append(sentRating)
    elif restName=="Subway":
        subratingList.append(restRating)
        subsentimentReviewList.append(sentRating)

#Create the variables for the Avg sentiments and ratings
bksent=mean(bksentimentReviewList)
bkrate=mean(bkratingList)
subsent=mean(subsentimentReviewList)
subrate=mean(subratingList)

#Print Average Sentiment and Rating
print("Burger King Average Sentiment:", round(bksent,3))
print("Burger King Ratings Average:", round(bkrate,2))
print("Subway Average Sentiment:", round(subsent,3))
print("Subway Ratings Average:", round(subrate,2))

#Write Average Sentiment and Rating into a new excel file
sheet1.cell(column=1, row=1).value="Resturant Name"
sheet1.cell(column=1, row=2).value="Burger King"
sheet1.cell(column=1, row=3).value="Subway"
sheet1.cell(column=2, row=1).value="Average Rating"
sheet1.cell(column=2, row=2).value=round(bkrate,2)
sheet1.cell(column=2, row=3).value=round(subrate,2)
sheet1.cell(column=3, row=1).value="Average Sentiment"
sheet1.cell(column=3, row=2).value=round(bksent,3)
sheet1.cell(column=3, row=3).value=round(subsent,3)

#Save into new book
restBook.save('Restaurant Summary.csv')
        
        
        
    

